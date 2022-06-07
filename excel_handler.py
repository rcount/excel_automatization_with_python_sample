import time

import pandas as pd
from openpyxl import load_workbook, Workbook
import xlwings as xw

from data_processor import get_data_of_player_by_position

def paste_data_in(sheet:Workbook, data):
    column = 1
    row = 2
    title_column = ''
    while title_column is not None:
        title_column = sheet.cell(row, column).value

        if title_column is None:
            continue

        row = 3
        for value in data[title_column]:
            if pd.isna(value):
                value = 0
            sheet.cell(row, column, value)
            row += 1
        row = 2
        column += 1


strikers, midfielders, defenders, goalkeepers = get_data_of_player_by_position()

spreadsheet = load_workbook(filename='the_ideal_lineup_template.xlsx')

paste_data_in(spreadsheet['Strikers'], strikers)
paste_data_in(spreadsheet['Midfielders'], midfielders)
paste_data_in(spreadsheet['Defenders'], defenders)
paste_data_in(spreadsheet['Goalkeepers'], goalkeepers)

FINAL_EXCEL = 'the_ideal_lineup_template_final.xlsx'
spreadsheet.save(FINAL_EXCEL)






